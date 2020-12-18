# *** coding: utf-8 ***
#@Time   : 2020/11/28 4:32 下午
#@Author : xueqing.wu
#@Email  : wuxueqing@126.com
#@File   : readData.py

import openpyxl
import yaml
import csv
from settings import DIR_NAME

class ReadData():
    def get_excel_list(self, filename):
        '''
        获取Excel数据，以[[],[],[]]形式显示
        :param filename: Excel文件地址
        :return: 所有测试用例
        '''
        wb = openpyxl.load_workbook(DIR_NAME + '/test_data/%s' % filename)
        ws = wb['测试用例']
        all_cases = []
        # 获取有效数据区域
        select_data_area = ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col= ws.max_column)
        for rows in select_data_area:
            case_list = [cell.value for cell in rows]
            all_cases.append(case_list)
        return all_cases

    def get_excel_dict(self, header, filename):
        '''
        获取Excel数据，以[{},{},{}]形式显示
        :param header: 表头
        :param filename: Excel文件地址
        :return: 所有测试用例
        '''
        wb = openpyxl.load_workbook(DIR_NAME + '/test_data/%s' % filename)
        ws = wb['测试用例']
        all_cases = []
        for row in range(2, ws.max_row+1):
            case_dict = {}
            col = 1
            for key in header:
                # 通过坐标获取每个单元格的内容
                case_dict[key] = ws.cell(row, col).value
                col += 1
            all_cases.append(case_dict)
        return all_cases

    def get_yaml(self, key, filename):
        '''
        获取yaml文件数据
        :param key: 某个用例
        :param filename: yaml文件地址
        :return: 所有测试用例
        '''
        with open(DIR_NAME + '/test_data/%s' % filename, 'r', encoding='utf-8') as f:
            yaml_data = yaml.safe_load(f)
            data_dict = yaml_data.get(key)
            case_object = data_dict.values()
            return list(case_object)

    def get_csv(self, filename):
        '''
        获取CSV文件数据
        :param filename: CSV文件地址
        :return: 所有测试用例
        '''
        with open(DIR_NAME + '/test_data/%s' % filename, 'r', encoding='utf-8') as f:
            # 跳过第一行数据
            next(f)
            all_cases = list(csv.reader(f))
            return all_cases


if __name__ == '__main__':
    print(ReadData().get_excel_list('login_data.xlsx'))
    print(ReadData().get_excel_dict(['username', 'password', 'exp', 'ids'], 'login_data.xlsx'))
    print(ReadData().get_yaml('test_login', 'login_data.yaml'))
    print(ReadData().get_csv('login_data.csv'))